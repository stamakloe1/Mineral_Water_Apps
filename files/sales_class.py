from tkinter import *
import datetime
from tkinter import messagebox
import re
import itertools
import pandas as pd, datetime as dt
from openpyxl import *
import numpy as np

root = Tk()

class sales_form:
    _params = [] # Protected Member can be accessed  within class or subclass
    # Private Members of the class
    __ck = StringVar()
    __c2 = StringVar()    #get date year
    __c3 = StringVar()    #get date month
    __c4 = StringVar()    #get date day
    __no_bgs = Entry()    #Number of bags
    __sl_bgs = Entry()    #bag sold
    __db_bgs = Entry()    #BAGS WENT TO DEPOT
    __dbt_bg = Entry()    #Bags in dept
    __ch_bgs = Entry()    #chopmoney bags
    __bn_bgs = Entry()    #Bonus bag
    __rtn_bg = Entry()    #Return bags
    __P_bags = Entry()    #Price per bag
    __dbprbg = Entry()
    __fuel_c = Entry()    #fuel cost
    __expens = Entry()    #expenses
    __sales = IntVar()    #Total bag sold
    __depot = IntVar()    #Total bag sold
    __debt_ = IntVar()    #Total bags in debt
    __chop_ = IntVar()    #total chop for workers
    __amunt = IntVar()    #Total Amount 
    
    
    #----------------------------------------------
    
    def __init__(self):
        self._params = ['12345','12345','1996-12-12','1996','12','12']
        
    def __del__(self):
        self._params = []

    def num_match(self,strg, search=re.compile(r'[^0-9.]').search):

        return not bool(search(strg))

    def mainget(self): 
        code  =  self.__ck.get()
        YeaR  =  self.__c2.get()
        MontH =  self.__c3.get()
        DatE  =  self.__c4.get()        
        nbgs = self.__no_bgs.get()
        slbg = self.__sl_bgs.get()
        dbbg = self.__db_bgs.get()
        dtbg = self.__dbt_bg.get()
        chbg = self.__ch_bgs.get()
        bnbg = self.__bn_bgs.get()
        rtbg = self.__rtn_bg.get()
        prbg = self.__P_bags.get()
        dbpr = self.__dbprbg.get()
        sale = self.__sales.get()
        depo = self.__depot.get()
        debt = self.__debt_.get()
        fuel = self.__fuel_c.get()
        chop = self.__chop_.get()
        exps = self.__expens.get()
        amnt = self.__amunt.get()
        
        #validate the forms
        date_format = '%d-%m-%y'
        flag=0
        try:
           # date_string = DatE + '-' + MontH + '-' + YeaR
            date_string = YeaR + '-' + MontH + '-' + DatE
            datetime.datetime.strptime(date_string, date_format)
        except ValueError:
            flag=1      
            
        ch = messagebox.askyesno("Submit", "Do you Want to Submit?")
        if ch==True:
            root.destroy()        
        
        self._params = [date_string,DatE,MontH,YeaR,nbgs,slbg,dbbg,dtbg,chbg,bnbg,rtbg,dbpr,prbg,sale,debt,depo,fuel,chop,exps,amnt]
        #self._params = self._params[:-]        
    def calculate(self):
            
        ch = messagebox.askyesno("Submit", "Do you Want to Submit?")
        if ch==True:
            root.destroy()        
        
        #self._params = [date_string,DatE,MontH,YeaR,nbgs,slbg,dbbg,dtbg,chbg,bnbg,rtbg,prbg,sale,debt,depo,fuel,chop,exps,amnt]        
        
        
        df = pd.read_excel("omft.xlsx", header=0)
        
        #df = df.apply(lambda row: row['Date'] - dt.timedelta(days=row['Date'].weekday()))
        #perweek = df.groupby(df['Date']).count()
        
        #print(perweek)
        
        #lst_row = ['Total'] + list(df.sum())[1:]
        #df2 = pd.DataFrame(data=[lst_row], columns=df.columns)
        #df = df.append(df2, ignore_index=True) 
        #df.loc['c_total']=df.sum(numeric_only=True, axis=0)
        lst_row = ['Total'] + list(df.sum())[1:]
        df2 = pd.DataFrame(data=[lst_row], columns=df.columns)
        #df2.style.set_properties(**{'background-color': 'black',
         #                  'color': 'green'})

        df = df.append(df2, ignore_index=True) 
        
        df.to_excel("omft.xlsx", index=False)
        return df
    
    def layout(self):

        root.title('GMC Daily_Sales')
        self.pic = PhotoImage(file="background.png")
        ba__ckground_label = Label(image=self.pic)
        ba__ckground_label.pack()
        ba__ckground_label.place(x=0, y=0, relwidth=1, relheight=1)
        ba__ckground_label.image = self.pic
        canvas = Canvas(root,width=1000,height=750,bg="grey")
        box=canvas.create_rectangle(1000,750,13,20,fill="snow3")
        canvas.pack(expand=YES)
        
        label=Label(canvas,text="NATURAL MINERAL WATER",font=("Times",30,"bold"),fg="brown4",bg="snow3",)
        label.place(x=150,y=30)
        label=Label(canvas,text="Daily_Sales",font=("Times",30,"bold"),fg="brown4",bg="snow3",)
        label.place(x=380,y=90)
        #Number of bag entry
        nobg=Label(canvas,text="Bags Qty:",font=("Times",14),fg="black",bg="snow3",)
        nobg.place(x=520,y=190) 
        self.__no_bgs= Entry(canvas,bg="white",bd=4)
        self.__no_bgs.place(x=615,y=190) 
        #bag sold entry
        bgsld = Label(canvas,text="Bag sold:",font=("Times",14),fg="black",bg="snow3",)
        bgsld.place(x=60,y=260)
        self.__sl_bgs= Entry(canvas,bg="white",bd=4)
        self.__sl_bgs.place(x=180,y=260)   
        #Bags in depot
        dpbag = Label(canvas,text="Depot Bag:",font=("Times",14),fg="black",bg="snow3",)
        dpbag.place(x=350,y=260)
        self.__db_bgs= Entry(canvas,bg="white",bd=4)
        self.__db_bgs.place(x=470,y=260)
        #Bags in dept
        dtbag = Label(canvas,text="Bag Debt:",font=("Times",14),fg="black",bg="snow3",)
        dtbag.place(x=650,y=260)
        self.__dbt_bg= Entry(canvas,bg="white",bd=4)
        self.__dbt_bg.place(x=750,y=260) 
        #bags for chop
        chbag = Label(canvas,text="chop_money:",font=("Times",14),fg="black",bg="snow3",)
        chbag.place(x=60,y=320)
        self.__ch_bgs= Entry(canvas,bg="white",bd=4)
        self.__ch_bgs.place(x=180,y=320)
        #bonus bag
        bnbag = Label(canvas,text="Bonus Bag:",font=("Times",14),fg="black",bg="snow3",)
        bnbag.place(x=350,y=320)
        self.__bn_bgs= Entry(canvas,bg="white",bd=4)
        self.__bn_bgs.place(x=470,y=320) 
        
        #numner of dep price
        dbp = Label(canvas,text="Depot Price:",font=("Times",14),fg="black",bg="snow3",)
        dbp.place(x=650,y=320)
        self.__rtn_bg= Entry(canvas,bg="white",bd=4)
        self.__rtn_bg.place(x=750,y=320)     
        #Price per bag
        prbag = Label(canvas,text="Bag Price:",font=("Times",14),fg="black",bg="snow3",)
        prbag.place(x=60,y=380)
        self.__P_bags= Entry(canvas,bg="white",bd=4)
        self.__P_bags.place(x=180,y=380)
                
        #Fuel
        gas = Label(canvas,text="Fuel Cost:",font=("Times",14),fg="black",bg="snow3",)
        gas.place(x=350,y=380)
        self.__fuel_c= Entry(canvas,bg="white",bd=4)
        self.__fuel_c.place(x=470,y=380) 
        #expenses
        expss = Label(canvas,text="Expenses",font=("Times",14),fg="black",bg="snow3",)
        expss.place(x=650,y=380)
        self.__expens= Entry(canvas,bg="white",bd=4)
        self.__expens.place(x=750,y=380)        
        
        #Date
        date=Label(canvas,text="Date:",font=("Times",14),fg="black",bg="snow3")
        date.place(x=60,y=190)     
        #date items day-mount-year
        list2 = []
        for i in range(2015,2050):
            list2.append('{}'.format(i))

        droplist2=OptionMenu(canvas,self.__c2,*list2)
        self.__c2.set("Year")
        droplist2.place(x=400,y=190)

        list3 = []
        for i in range(1,13):
            list3.append('{}'.format(i))

        droplist3=OptionMenu(canvas,self.__c3,*list3)
        self.__c3.set("Month")
        droplist3.place(x=280,y=190)

        list4 = []
        for i in range(1,32):
            list4.append('{}'.format(i))

        droplist4=OptionMenu(canvas,self.__c4,*list4)
        self.__c4.set("Date")
        droplist4.place(x=180,y=190)        
        
        #Submit button
        submt = Button(canvas,text="Submit",width=20,height=2,bd=4,font=("Times",10,"bold"),bg="lightblue",command=self.mainget)
        submt.place(x=300,y=500)   
        
        Cal = Button(canvas,text="Calculate",width=20,height=2,bd=4,font=("Times",10,"bold"),bg="light green",command=self.calculate)
        Cal.place(x=600,y=500)        
 
#class to insert and staore excel file      
class Store_data(sales_form): # Derived class of Super class sales_form (Inheritance)
    def __init__(self,sheet):
        sales_form.__init__(self)
        self.sheet = sheet

    def __del__(self):
        sales_form.__del__(self)

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet
        self.sheet.column_dimensions['A'].width = 10
        self.sheet.column_dimensions['B'].width = 13
        self.sheet.column_dimensions['C'].width = 13
        self.sheet.column_dimensions['D'].width = 15
        self.sheet.column_dimensions['E'].width = 12
        self.sheet.column_dimensions['F'].width = 10
        self.sheet.column_dimensions['G'].width = 10
        self.sheet.column_dimensions['H'].width = 10
        self.sheet.column_dimensions['I'].width = 10
        self.sheet.column_dimensions['J'].width = 8
        self.sheet.column_dimensions['K'].width = 15
        self.sheet.column_dimensions['L'].width = 15
        self.sheet.column_dimensions['M'].width = 15
        self.sheet.column_dimensions['N'].width = 12
        self.sheet.column_dimensions['O'].width = 12
        self.sheet.column_dimensions['P'].width = 12  
        self.sheet.column_dimensions['Q'].width = 15  

        # write given data to an excel spreadsheet
        # at particular location
        self.sheet.cell(row=1, column=1).value = "Date"
        self.sheet.cell(row=1, column=2).value = "NO.Bags"
        self.sheet.cell(row=1, column=3).value = "Bag Sale"
        self.sheet.cell(row=1, column=4).value = "Depo bag"
        self.sheet.cell(row=1, column=5).value = "Debt bag"
        self.sheet.cell(row=1, column=6).value = "chpBag"
        self.sheet.cell(row=1, column=7).value = "Bonus"
        self.sheet.cell(row=1, column=8).value = "Returns"
        self.sheet.cell(row=1, column=9).value = "Depo_Price"
        self.sheet.cell(row=1, column=10).value = "Price"
        self.sheet.cell(row=1, column=11).value = "Total Sales"
        self.sheet.cell(row=1, column=12).value = "Total Debt"
        self.sheet.cell(row=1, column=13).value = "Total Depot"
        self.sheet.cell(row=1, column=14).value = "Fuel"
        self.sheet.cell(row=1, column=15).value = "chpMoney"
        self.sheet.cell(row=1, column=16).value = "Expenses" 
        self.sheet.cell(row=1, column=17).value = "Total Amount"



    def insert(self):

        current_row = self.sheet.max_row
        current_column = self.sheet.max_column
        self.sheet.cell(row=current_row + 1, column=1).value = self._params[0]
        self.sheet.cell(row=current_row + 1, column=2).value = self._params[4]
        self.sheet.cell(row=current_row + 1, column=3).value = self._params[5]
        self.sheet.cell(row=current_row + 1, column=4).value = self._params[6]        
        self.sheet.cell(row=current_row + 1, column=5).value = self._params[7]
        self.sheet.cell(row=current_row + 1, column=6).value = self._params[8]
        self.sheet.cell(row=current_row + 1, column=7).value = self._params[9]
        self.sheet.cell(row=current_row + 1, column=9).value = self._params[10]
        self.sheet.cell(row=current_row + 1, column=10).value = self._params[12]
        self.sheet.cell(row=current_row + 1, column=14).value = self._params[16]
        self.sheet.cell(row=current_row + 1, column=16).value = self._params[18]
        #self.sheet.cell(row=current_row + 1, column=17).value = self._params[19]
        #self.sheet.cell(row=current_row + 1, column=12).value = self._params[14]
        #self.sheet.cell(row=current_row + 1, column=13).value = self._params[15]
        #self.sheet.cell(row=current_row + 1, column=14).value = self._params[16]        
      
    
def calculate():
    df = pd.read_excel("omft.xlsx")
    df['Total Sales'] = None
    df['Total  Debt'] = None
    df['Total Depot'] = None 
    df['chpMoney'] = None
    df['Total Amount'] = None
    df['Returns'] = None
    
    index_bgsale = df.columns.get_loc('Bag Sale')
    index_price = df.columns.get_loc('Price')
    index_sales = df.columns.get_loc('Total Sales')
    index_dpbag = df.columns.get_loc('Depo bag')
    index_tdepo = df.columns.get_loc('Total Depot')
    index_tdbt = df.columns.get_loc('Total Debt')
    index_chmny = df.columns.get_loc('chpMoney')
    index_chbg = df.columns.get_loc('chpBag')    
    index_chmny = df.columns.get_loc('chpMoney')
    index_ttamnt = df.columns.get_loc('Total Amount')     
    index_dtbag = df.columns.get_loc('Debt bag')
    index_rtbag = df.columns.get_loc('Returns')
    index_gas = df.columns.get_loc('Fuel')
    index_exp = df.columns.get_loc('Expenses')
    index_nbags = df.columns.get_loc('NO.Bags')
    
    
    for row in range(0, len(df)):
        #df.iat[row, index_rtbag] = df.iat[row,
         #                                  index_nbags] - df.iat[row, index_bgsale]- df.iat[row, 
          #                                                                                  index_dtbag] - df.iat[row, index_dpbag]         
        df.iat[row, index_sales] = df.iat[row,
                                         index_bgsale] * df.iat[row, index_price] 
        df.iat[row, index_tdbt] = df.iat[row,
                                           index_dtbag] * df.iat[row, index_price]         
        df.iat[row, index_tdepo] = df.iat[row,
                                          index_dpbag] * df.iat[row, index_price]               
        df.iat[row, index_chmny] = df.iat[row,
                                           index_chbg] * df.iat[row, index_price]
        df.iat[row, index_ttamnt] = df.iat[row,
                                           index_sales] + df.iat[row, index_tdepo]- df.iat[row, 
                                                                                           index_gas] - df.iat[row, index_exp]        
      
    #lst_row = ['Total'] + list(df.sum())[1:]
    #df2 = pd.DataFrame(data=[lst_row], columns=df.columns)
    #df = df.append(df2, ignore_index=True)
    #df.groupby('Date','NO.Bags','Bag Sale','Depo bag','Debt bag','chpBag','Bonus','Returns',
                #'Price', 'Total Sales','Total Debt','Total Depot','Fuel','chpMoney','Expenses','Total Amount').sum()
    #df.sum()
    
    
    #df.drop(["Total Debt"], axis = 1, inplace = True)
    #df = df.iloc[17:]
        
    df.to_excel("omft.xlsx", index=False)            

def main():
    try:
        wb = load_workbook('omft.xlsx')
        sheet = wb.active
        obj1 = Store_data(sheet)
        obj1.layout()
        root.mainloop()
        obj1.excel()
        obj1.insert()
        wb.save('omft.xlsx')
        calculate()
        try:
            #calculate()
            wb = load_workbook('omft.xlsx')
            sheet = wb.active   
            obj1 = Store_data(sheet)
            obj1.excel()
            #worksheet = sheet.getWorksheets().get(0)
            #cell = worksheet.getCells()
            #cells.groupRows(0,7,True)
            ##cells.groupColumns(1,16,True)
            sheet.delete_cols(18)
            wb.save('omft.xlsx')
        except:
            pass
            #calculate()
        
        
    except PermissionError:
        print('Permission denied : excel.xlsx file is already opened for other use')
        
if __name__ == "__main__":
    main()
else:
    pass      
        
        